# Copyright (c) 2023-2023 Huawei Technologies Co., Ltd.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

import os
import shutil
import pandas as pd
import openpyxl

from msquickcmp.cmp_process import csv_sum


def test_csv_sum():
    if os.path.exists('./test_resource/test_csv_sum'):
        shutil.rmtree('./test_resource/test_csv_sum')
    
    os.mkdir('./test_resource/test_csv_sum')
    os.mkdir('./test_resource/test_csv_sum/2023072009')
    os.mkdir('./test_resource/test_csv_sum/2023072009/images-2_3_638_640')
    os.mkdir('./test_resource/test_csv_sum/2023072009/images-2_3_640_640')

    df1 = pd.DataFrame({'A': [1, 2, 3], 'B': ['a', 'b', 'c']})
    df1.to_csv('./test_resource/test_csv_sum/2023072009/images-2_3_638_640/file1.csv', index=False)
    
    df2 = pd.DataFrame({'A': [1, 2, 3], 'B': ['b', 'c', 'd']})
    df2.to_csv('./test_resource/test_csv_sum/2023072009/images-2_3_640_640/file2.csv', index=False)


    writer = pd.ExcelWriter("./test_resource/test_csv_sum/expected_output.xlsx", engine='xlsxwriter')
    df1.to_excel(writer, sheet_name='images-2_3_638_640', index=False)
    df2.to_excel(writer, sheet_name='images-2_3_640_640', index=False)
    writer.save()

    csv_sum("./test_resource/test_csv_sum/2023072009")

    result_summary = openpyxl.load_workbook('./test_resource/test_csv_sum/2023072009/result_summary.xlsx')
    expected_output = openpyxl.load_workbook('./test_resource/test_csv_sum/expected_output.xlsx')

    sheets1 = result_summary.sheetnames
    sheets2 = expected_output.sheetnames
    
    assert len(sheets1) == len(sheets2)

    for sheet_name in sheets1:
        sheet1 = result_summary[sheet_name]
        sheet2 = expected_output[sheet_name]

        assert sheet1.max_row == sheet2.max_row
        assert sheet1.max_column == sheet2.max_column
  
        for row in range(1, sheet1.max_row + 1):
            for col in range(1, sheet1.max_column + 1):
                assert sheet1.cell(row=row, column=col).value == sheet2.cell(row=row, column=col).value

    shutil.rmtree('./test_resource/test_csv_sum')